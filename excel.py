"""Render a Report into a styled multi-sheet workbook."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

import config
import jalali
from report import Report

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F3864")
LABEL_FONT = Font(bold=True, color="404040")
STRIPE_FILL = PatternFill("solid", fgColor="F2F5FA")
WEEKEND_FILL = PatternFill("solid", fgColor="FDF3F3")

_THIN = Side(style="thin", color="D0D7E5")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

TOP_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center")


class Column:
    """A worksheet column: header text, width, and how to pull its value."""

    def __init__(
        self,
        title: str,
        width: int,
        getter=None,
        *,
        numeric: bool = False,
        key: str | None = None,
    ):
        self.title = title
        self.width = width
        self.getter = getter
        self.numeric = numeric
        # Stable lowercase handle used by --hide, so renaming a title on the
        # sheet never breaks a caller's flags.
        self.key = key or title.lower().replace(" ", "-").split("(")[0].strip("- ")


def visible(columns: list[Column], hidden: set[str]) -> list[Column]:
    return [column for column in columns if column.key not in hidden]


def _write_header(sheet: Worksheet, columns: list[Column], row: int = 1) -> None:
    for index, column in enumerate(columns, start=1):
        cell = sheet.cell(row=row, column=index, value=column.title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
        sheet.column_dimensions[get_column_letter(index)].width = column.width
    sheet.freeze_panes = sheet.cell(row=row + 1, column=1)


def _write_rows(sheet: Worksheet, columns: list[Column], items, *, start: int = 2) -> None:
    for offset, item in enumerate(items):
        row = start + offset
        for index, column in enumerate(columns, start=1):
            cell = sheet.cell(row=row, column=index, value=column.getter(item))
            cell.alignment = CENTER if column.numeric else TOP_LEFT
            cell.border = BORDER
            if offset % 2:
                cell.fill = STRIPE_FILL


def _autofilter(sheet: Worksheet, columns: list[Column], rows: int) -> None:
    if rows:
        last = get_column_letter(len(columns))
        sheet.auto_filter.ref = f"A1:{last}{rows + 1}"


def _badge(sheet: Worksheet, row: int, columns: int = 4) -> None:
    """Attribution footer, shown once at the bottom of each summary sheet."""
    cell = sheet.cell(row=row, column=1, value=f"{config.COPYRIGHT} - Git Activity Report")
    cell.font = Font(size=9, color="9CA3AF", italic=True)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=columns)


def _section(sheet: Worksheet, row: int, heading: str) -> int:
    for column in range(1, 5):
        cell = sheet.cell(row=row, column=column)
        cell.fill = HEADER_FILL
        if column == 1:
            cell.value = heading
            cell.font = HEADER_FONT
    return row + 1


def _sheet_name(prefix: str, title: str) -> str:
    """Excel caps sheet names at 31 chars and rejects []:*?/\\ ."""
    name = f"{prefix} - {title}" if prefix else title
    for bad in "[]:*?/\\":
        name = name.replace(bad, "-")
    if len(name) > 31:
        keep = 31 - len(title) - 3
        name = f"{prefix[:max(keep, 1)]}... {title}"[:31] if keep < 1 else f"{prefix[:keep]} - {title}"
    return name


def _summary_sheet(book: Workbook, report: Report, prefix: str = "") -> None:
    sheet = book.create_sheet(_sheet_name(prefix, "Summary"))
    sheet.sheet_view.showGridLines = False
    for column, width in zip("ABCD", (30, 34, 12, 26)):
        sheet.column_dimensions[column].width = width

    sheet["A1"] = f"{report.project} - Development Report"
    sheet["A1"].font = TITLE_FONT
    sheet.merge_cells("A1:D1")

    subtitle = sheet.cell(row=2, column=1, value=f"Period: {report.label} (Jalali)")
    subtitle.font = MUTED_FONT
    sheet.merge_cells("A2:D2")

    row = 4
    facts = [
        ("Author", "; ".join(name for name, _ in report.authors) or "-"),
        ("Total commits", len(report.commits)),
        ("Active days", report.active_days),
        ("Files changed", report.total_files),
        ("Lines added", report.total_insertions),
        ("Lines removed", report.total_deletions),
    ]
    if report.generated_insertions:
        facts.append(
            (
                "Generated code",
                f"{report.generated_insertions:,} lines in {report.generated_files} "
                f"files (excluded from totals above)",
            )
        )
    for label, value in facts:
        sheet.cell(row=row, column=1, value=label).font = LABEL_FONT
        sheet.cell(row=row, column=2, value=value).alignment = TOP_LEFT
        row += 1

    if report.scope_breakdown:
        row = _section(sheet, row + 1, "What Was Built (by area)")
        busiest = max(count for _, count in report.scope_breakdown)
        for name, count in report.scope_breakdown:
            sheet.cell(row=row, column=1, value=name).alignment = TOP_LEFT
            sheet.cell(row=row, column=2, value="=" * max(1, round(count / busiest * 24)))
            sheet.cell(row=row, column=2).font = BAR_FONT
            sheet.cell(row=row, column=3, value=count).alignment = CENTER
            row += 1

    if report.type_breakdown:
        row = _section(sheet, row + 1, "Type of Work")
        total = len(report.commits)
        for name, count in report.type_breakdown:
            sheet.cell(row=row, column=1, value=_TYPE_LABELS.get(name, name)).alignment = TOP_LEFT
            sheet.cell(row=row, column=2, value="=" * max(1, round(count / total * 24)))
            sheet.cell(row=row, column=2).font = BAR_FONT
            sheet.cell(row=row, column=3, value=count).alignment = CENTER
            sheet.cell(row=row, column=4, value=f"{count / total:.0%}").alignment = CENTER
            row += 1

    if report.days:
        row = _section(sheet, row + 1, "Daily Activity")
        busiest = max(len(day.commits) for day in report.days)
        for day in report.days:
            count = len(day.commits)
            sheet.cell(row=row, column=1, value=f"{day.jalali_date}  {day.weekday}").alignment = TOP_LEFT
            sheet.cell(row=row, column=2, value="=" * max(1, round(count / busiest * 24)))
            sheet.cell(row=row, column=2).font = BAR_FONT
            sheet.cell(row=row, column=3, value=count).alignment = CENTER
            sheet.cell(row=row, column=4, value=f"+{day.insertions} / -{day.deletions}").alignment = CENTER
            row += 1

    if report.days:
        row = _section(sheet, row + 1, "Commit Titles by Day")
        for day in report.days:
            label = sheet.cell(row=row, column=1, value=f"{day.jalali_date}  {day.weekday}")
            label.font = GROUP_FONT
            label.fill = GROUP_FILL
            label.alignment = TOP_LEFT
            for column in range(2, 5):
                sheet.cell(row=row, column=column).fill = GROUP_FILL
            row += 1
            for commit in day.commits:
                cell = sheet.cell(row=row, column=1, value=f"    {commit.subject}")
                cell.alignment = TOP_LEFT
                sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
                sheet.row_dimensions[row].outlineLevel = 1
                sheet.row_dimensions[row].hidden = True
                row += 1
        sheet.sheet_properties.outlinePr.summaryBelow = False

    _badge(sheet, row + 1)


def _daily_sheet(book: Workbook, report: Report, hidden: set[str], prefix: str = "") -> None:
    sheet = book.create_sheet(_sheet_name(prefix, "Daily Task List"))
    columns = visible(
        [
            Column("Date", 14, lambda d: d.jalali_date),
            Column("Weekday", 12, lambda d: d.weekday),
            Column("Month", 13, lambda d: d.month),
            Column("Commits", 9, lambda d: len(d.commits), numeric=True),
            Column("First", 8, lambda d: d.first_commit, numeric=True),
            Column("Last", 8, lambda d: d.last_commit, numeric=True),
            Column("Span (h)", 9, lambda d: d.span_hours, numeric=True, key="span"),
            Column("Files", 8, lambda d: d.files_changed, numeric=True),
            Column("Added", 9, lambda d: d.insertions, numeric=True),
            Column("Removed", 9, lambda d: d.deletions, numeric=True),
            Column("Scopes", 22, lambda d: d.scopes),
            Column("Types", 18, lambda d: d.types),
            Column("Tasks Completed", 80, lambda d: d.tasks, key="tasks"),
            Column("Commit Titles", 80, lambda d: d.commit_titles, key="titles"),
            Column(
                "Commit Descriptions",
                90,
                lambda d: d.commit_descriptions,
                key="descriptions",
            ),
        ],
        hidden,
    )
    _write_header(sheet, columns)
    _write_rows(sheet, columns, report.days)

    for offset, entry in enumerate(report.days):
        if entry.weekday in ("Thursday", "Friday"):
            for index in range(1, len(columns) + 1):
                sheet.cell(row=2 + offset, column=index).fill = WEEKEND_FILL

    _autofilter(sheet, columns, len(report.days))


def _commit_sheet(book: Workbook, report: Report, hidden: set[str], prefix: str = "") -> None:
    sheet = book.create_sheet(_sheet_name(prefix, "Commits"))
    columns = visible(
        [
            Column("Date", 14, lambda c: jalali.format_date(c.authored_at)),
            Column("Weekday", 12, lambda c: jalali.weekday_name(c.authored_at)),
            Column("Time", 9, lambda c: c.authored_at.strftime("%H:%M"), numeric=True),
            Column("Author", 18, lambda c: c.author_name),
            Column("Type", 11, lambda c: c.commit_type),
            Column("Scope", 18, lambda c: c.scope),
            Column("Commit Title", 74, lambda c: c.subject, key="title"),
            Column("Commit Description", 90, lambda c: c.body, key="descriptions"),
            Column("Files", 8, lambda c: c.files_changed, numeric=True),
            Column("Added", 9, lambda c: c.insertions, numeric=True),
            Column("Removed", 9, lambda c: c.deletions, numeric=True),
            Column("Commit", 12, lambda c: c.short_hash, key="hash"),
        ],
        hidden,
    )
    _write_header(sheet, columns)
    _write_rows(sheet, columns, report.commits)
    _autofilter(sheet, columns, len(report.commits))


GROUP_FILL = PatternFill("solid", fgColor="E8EDF7")
GROUP_FONT = Font(bold=True, color="1F3864")
MUTED_FONT = Font(color="6B7280")

# Inline bars: a monospace font keeps every '=' the same width, so the bars
# line up as a chart without needing an embedded chart object.
BAR_FONT = Font(name="Consolas", color="4472C4")

_TYPE_LABELS = {
    "feat": "New features",
    "fix": "Bug fixes",
    "refactor": "Code improvements",
    "perf": "Performance",
    "style": "Styling",
    "docs": "Documentation",
    "test": "Tests",
    "chore": "Maintenance",
    "build": "Build system",
    "ci": "CI pipeline",
    "revert": "Reverts",
    "other": "Other",
}


def _files_sheet(book: Workbook, report: Report, prefix: str = "") -> None:
    """One bold row per commit, with its files as a collapsible child group."""
    sheet = book.create_sheet(_sheet_name(prefix, "Changed Files"))
    columns = [
        Column("Date", 14, None),
        Column("Change / File", 96, None),
        Column("Area", 22, None),
        Column("Added", 9, None, numeric=True),
        Column("Removed", 9, None, numeric=True),
    ]
    _write_header(sheet, columns)

    row = 2
    for commit in report.commits:
        header = [
            jalali.format_date(commit.authored_at),
            commit.summary,
            commit.scope or commit.commit_type,
            commit.insertions,
            commit.deletions,
        ]
        for index, value in enumerate(header, start=1):
            cell = sheet.cell(row=row, column=index, value=value)
            cell.fill = GROUP_FILL
            cell.font = GROUP_FONT
            cell.border = BORDER
            cell.alignment = CENTER if index >= 4 else TOP_LEFT
        row += 1

        for path in commit.files:
            sheet.cell(row=row, column=1)
            cell = sheet.cell(row=row, column=2, value=f"    {path}")
            cell.alignment = TOP_LEFT
            cell.font = MUTED_FONT
            for index in range(1, len(columns) + 1):
                sheet.cell(row=row, column=index).border = BORDER
            # outlineLevel makes Excel draw the [+]/[-] control in the gutter.
            sheet.row_dimensions[row].outlineLevel = 1
            sheet.row_dimensions[row].hidden = True
            row += 1

    sheet.sheet_properties.outlinePr.summaryBelow = False


def _overview_sheet(book: Workbook, reports: list[tuple[str, Report]], period: str) -> None:
    """Cross-project roll-up, shown first when more than one project is reported."""
    sheet = book.create_sheet("Overview", 0)
    sheet.sheet_view.showGridLines = False
    for column, width in zip("ABCDEF", (30, 28, 12, 12, 14, 14)):
        sheet.column_dimensions[column].width = width

    sheet["A1"] = "Development Report - All Projects"
    sheet["A1"].font = TITLE_FONT
    sheet.merge_cells("A1:F1")
    subtitle = sheet.cell(row=2, column=1, value=f"Period: {period} (Jalali)")
    subtitle.font = MUTED_FONT
    sheet.merge_cells("A2:F2")

    totals = {
        "Projects": len(reports),
        "Total commits": sum(len(r.commits) for _, r in reports),
        "Active days": len({d.day for _, r in reports for d in r.days}),
        "Files changed": sum(r.total_files for _, r in reports),
        "Lines added": sum(r.total_insertions for _, r in reports),
        "Lines removed": sum(r.total_deletions for _, r in reports),
    }
    row = 4
    for label, value in totals.items():
        sheet.cell(row=row, column=1, value=label).font = LABEL_FONT
        sheet.cell(row=row, column=2, value=value).alignment = TOP_LEFT
        row += 1

    row = _section(sheet, row + 1, "By Project")
    headers = ["Project", "Activity", "Commits", "Days", "Added", "Removed"]
    for index, title in enumerate(headers, start=1):
        cell = sheet.cell(row=row, column=index, value=title)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER
    row += 1

    busiest = max((len(r.commits) for _, r in reports), default=1) or 1
    for name, report in sorted(reports, key=lambda item: -len(item[1].commits)):
        count = len(report.commits)
        values = [
            name,
            "=" * max(1, round(count / busiest * 22)),
            count,
            report.active_days,
            report.total_insertions,
            report.total_deletions,
        ]
        for index, value in enumerate(values, start=1):
            cell = sheet.cell(row=row, column=index, value=value)
            cell.border = BORDER
            cell.alignment = TOP_LEFT if index == 1 else CENTER
            if index == 2:
                cell.font = BAR_FONT
        row += 1

    _badge(sheet, row + 1, columns=6)


def write(report: Report, destination: Path, hidden: set[str] | None = None) -> Path:
    return write_multi([(report.project, report)], destination, hidden)


def write_multi(
    reports: list[tuple[str, Report]],
    destination: Path,
    hidden: set[str] | None = None,
) -> Path:
    """Write one workbook; multiple projects get an Overview sheet and prefixes."""
    if not reports:
        raise ValueError("nothing to write")

    book = Workbook()
    book.remove(book.active)
    hidden = hidden or set()
    multi = len(reports) > 1

    for index, (name, report) in enumerate(reports, start=1):
        prefix = ""
        if multi:
            # Long names truncate to 31 chars, so two similar projects can collide
            # and be silently renamed by Excel. Prefix an index when that happens.
            collides = any(
                _sheet_name(other, "Summary") == _sheet_name(name, "Summary")
                for other, _ in reports[: index - 1]
            )
            prefix = f"{index}.{name}" if collides else name
        _summary_sheet(book, report, prefix)
        _daily_sheet(book, report, hidden, prefix)
        _commit_sheet(book, report, hidden, prefix)
        _files_sheet(book, report, prefix)

    if multi:
        _overview_sheet(book, reports, reports[0][1].label)

    book.properties.creator = config.AUTHOR_TAG
    book.properties.title = "Git Activity Report"

    destination.parent.mkdir(parents=True, exist_ok=True)
    book.save(destination)
    return destination
